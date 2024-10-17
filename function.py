import sys, time
from py3270 import CommandError

from openpyxl import load_workbook

def string_wait(em_v,string):
    while True:
        em_v.wait_for_field()
        c = em_v.exec_command(b"Ascii()")
        if string in b"\n".join(c.data).decode():
            return

def excel_data_load(excelf):
    # Function to read data from the Excel sheet, process it, and return the record list
    wb = load_workbook(excelf)

    # Select the main sheet and the backup sheet
    sheet_to_read = wb['MAIN']  
    backup_sheet = wb['BackupSheet']  

    record = []

    # Read headers from the first row of the main sheet
    headers = [cell.value for cell in sheet_to_read[1]]

    # Format cell values as string
    def format_cell_value(cell_value):
        if cell_value is None or cell_value == '':  
            return ' '  
        elif isinstance(cell_value, (int, float)):  
            return f'{cell_value}'  
        return cell_value  

    # Read all the rows starting from the second row (assuming the first row is headers)
    for row in sheet_to_read.iter_rows(min_row=2, values_only=True):
        if not any(row):  # Stop reading if a row is completely empty
            break
        # Apply the formatting to each cell in the row
        formatted_row = [format_cell_value(cell) for cell in row]  # Use format_cell_value directly
        row_data = dict(zip(headers, formatted_row))
        record.append(row_data)  # Append the formatted dictionary to the list

    # Append the dictionary data (row by row) into the backup sheet
    for row_data in record:
        backup_sheet.append(list(row_data.values()))

    # Find the maximum row with data
    max_row = sheet_to_read.max_row

    # Deleting the rows from the main sheet (starting from row 2, keeping the header)
    #for row_num in range(2, max_row + 1):
    #    sheet_to_read.delete_rows(2)  

    
    # Save the workbook 
    wb.save(excelf)

    #print("Data has been successfully appended to the backup sheet and formatted as required!")
    #print(record)
    return record

def policy_add_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(24,10,'EA1 ',4)
        em_v.fill_field(24,20,'  ',2)
        em_v.fill_field(24,28,'T1',2)
        em_v.fill_field(24,38,ea_data['POL_NO'].strip().upper(),15)
        em_v.fill_field(24,61,' ',5)
        em_v.fill_field(24,70,ea_data['CO'].strip().upper(),3)
        em_v.fill_field(24,79,'N',1)

        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(22,3,38)
        error_2 = em_v.string_get(22,42,38)
        error_3 = em_v.string_get(23,3,38)
        error_4 = em_v.string_get(23,42,38)
        
        if em_v.string_get(2,31,18) == '*** WELCOME TO ***' and em_v.string_get(24,10,4) == '    ':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("ea1-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'any' key is missing from dict
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea1_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(3,15,ea_data['STAT'].strip().upper(),3)
        em_v.fill_field(3,43,ea_data['POL_NO'].strip().upper(),10)
        em_v.fill_field(5,21,ea_data['PLAN_CODE'].strip().upper(),6)
        em_v.fill_field(5,42,ea_data['NO_OF_UNITS'].strip().upper(),3)
        em_v.fill_field(6,17,ea_data['POL_EFF_DATE'].strip().upper(),8)
        em_v.fill_field(7,52,ea_data['LINE_OF_BUSINESS'].strip().upper(),3)
        em_v.fill_field(7,80,ea_data['OWNERSHIP_CODE'].strip().upper(),1)
        em_v.fill_field(10,37,ea_data['RES_ST'].strip().upper(),3)
        em_v.fill_field(10,53,ea_data['ISS_ST'].strip().upper(),3)        
        em_v.fill_field(11,17,ea_data['INS_LAST'].strip().upper(),15)
        em_v.fill_field(11,41,ea_data['INS_FIRST_NAME'].strip().upper(),20)
        em_v.fill_field(12,7,ea_data['INS_DOB'].strip().upper(),8)
        em_v.fill_field(12,34,ea_data['INS_Sex'].strip().upper(),1)
        em_v.fill_field(12,40,ea_data['INS_SSN'].strip().upper(),15)
        em_v.fill_field(12,67,ea_data['INS_SSN_Verify'].strip().upper(),1)
        em_v.fill_field(13,10,ea_data['INS_ADDR_1'].strip().upper(),30)
        em_v.fill_field(13,51,ea_data['INS_ADDR_2'].strip().upper(),30)
        em_v.fill_field(14,8,ea_data['INS_CITY'].strip().upper(),30)
        em_v.fill_field(14,43,ea_data['INS_ST'].strip().upper(),3)
        em_v.fill_field(14,55,ea_data['INS_ZIP_CD'].strip().upper(),9)
        em_v.fill_field(14,76,ea_data['INS_COUNTRY_CD'].strip().upper(),2)
        em_v.fill_field(15,18,ea_data['INS_EMAIL_ADDRESS'].strip().upper(),50)
        em_v.fill_field(16,9,ea_data['INS_PHONE'].strip().upper(),12)
 

        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(22,3,38)
        error_2 = em_v.string_get(22,42,38)
        error_3 = em_v.string_get(23,3,38)
        error_4 = em_v.string_get(23,42,38)
        
        if em_v.string_get(2,35,10) == 'POLICY ADD' and em_v.string_get(24,10,4) == 'EA2 ':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("ea1-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'any' key is missing from dict
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea2_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(10, 24, ea_data['MAIL_CODES_NFO_ORIG'].strip().upper(), 1)
        #em_v.fill_field(10, 24, ea_data['MAIL_CODES_NFO_DUPL'].strip().upper(), 1)
        em_v.fill_field(10, 46, ea_data['MAIL_CODES_OTHER_ORIG'].strip().upper(), 1)
        #em_v.fill_field(10, 24, ea_data['MAIL_CODES_OTHER_DUPL'].strip().upper(), 1)
        #em_v.fill_field(10, 24, ea_data['SIGNED_STATE'].strip().upper(), 1)
        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(19,3,38)
        error_2 = em_v.string_get(19,42,38)
        error_3 = em_v.string_get(20,3,38)
        error_4 = em_v.string_get(20,42,38)
        
        if em_v.string_get(3,25,31) == '*** PLAN INFORMATION (EA2) ***':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("ea2-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea3_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(4, 41, ea_data['BILL_OPT'].strip().upper(), 1)
        em_v.fill_field(4, 51, ea_data['BILL_FREQ'].strip().upper(), 2)
        em_v.fill_field(9, 60, ea_data['BILL_MAIL_CODE_ORIG'].strip().upper(), 1)
        #em_v.fill_field(9, 60, ea_data['PAC_TYPE'].strip().upper(), 1)
        #em_v.fill_field(9, 60, ea_data['BILL_CL_NO'].strip().upper(), 1)
        #em_v.fill_field(9, 60, ea_data['CHC'].strip().upper(), 1)
        #em_v.fill_field(9, 60, ea_data['BRC'].strip().upper(), 1)
        #em_v.fill_field(9, 60, ea_data['ACCOUNT'].strip().upper(), 1)
        
        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)


        error_1 = em_v.string_get(21,3,38)
        error_2 = em_v.string_get(21,42,38)
        error_3 = em_v.string_get(23,3,38)
        error_4 = em_v.string_get(23,42,38)
        
        if em_v.string_get(3,24,33) == '*** BILLING INFORMATION (EA3) ***':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("ea3-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea4_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(7, 7, ea_data['ROLE_TYPE'].strip().upper(), 1)
        em_v.fill_field(7, 10, ea_data['1YR_PCT'].strip().upper(), 3)
        em_v.fill_field(7, 17, ea_data['RNWL_PCT'].strip().upper(), 3)
        em_v.fill_field(7, 24, ea_data['COMM_AGENCY'].strip().upper(), 6)
        em_v.fill_field(7, 32, ea_data['PERSON_CODE'].strip().upper(), 10)
        em_v.fill_field(7, 44, ea_data['AGT_PRF'].strip().upper(), 3)
        em_v.fill_field(7, 49, ea_data['AGT_CONTR'].strip().upper(), 5)
        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        error_3 = em_v.string_get(24,3,38).strip()
        error_4 = em_v.string_get(24,42,38).strip()
        
        if em_v.string_get(3,24,35) == '*** PRODUCER INFORMATION (EA-4) ***' and (error_1 != 'C311 CANT ACCESS ACR'):   
            if error_1 != ' ' or error_2 != ' ' :  # Check for screen validation errors after submit
                print("ea4-screen-error" + error_1 + error_2 + error_3 + error_4 )
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea5_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data

        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        
        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,14,57) == '*** REINSURANCE/NON STANDARD RATING INFORMATION (EA5) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea5-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.") 
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea6_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
                
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,23,37) == '*** BENEFICIARY INFORMATION (EA6) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea6-screen-error"+"$$" + error_1 +"$$"+ error_2 +"$$")
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea7_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        
        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,23,37) == '*** OWNER/PAYOR INFORMATION (EA7) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea7-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea8_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
       


        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,20,45) == '*** ALTERNATE ADDRESSEE INFORMATION (EA8) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea8-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea9_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
                
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,24,31) == '*** REPLACEMENT INFORMATION ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea9-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea10_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
                
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,21,41) == '*** DIVIDEND PAYEE INFORMATION (EA10) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea3-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def ea11_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(9, 39, ea_data['REPL_IND'].strip().upper(), 1)


        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
        
        if em_v.string_get(3,25,31) == '*** USER INFORMATION (EA11) ***':   
            if error_1 != '' or error_2 != '' :  # Check for screen validation errors after submit
                print("ea11-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def rider_add_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(24,10,'GSB ',4)
        em_v.fill_field(24,20,'  ',2)
        em_v.fill_field(24,28,'T1',2)
        em_v.fill_field(24,38,ea_data['POL_NO'].strip().upper(),15)
        em_v.fill_field(24,61,' ',5)
        em_v.fill_field(24,70,ea_data['CO'].strip().upper(),3)
        em_v.fill_field(24,79,'N',1)

        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(22,3,38)
        error_2 = em_v.string_get(22,42,38)
        error_3 = em_v.string_get(23,3,38)
        error_4 = em_v.string_get(23,42,38)
        
        if em_v.string_get(2,31,18) == '*** WELCOME TO ***' and em_v.string_get(24,10,4) == '    ':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("gsb-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'any' key is missing from dict
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def GS_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(4, 10, ea_data['COV_IND'].strip().upper(), 2)
        em_v.fill_field(4, 46, ea_data['COV_TYPE'].strip().upper(), 5)
        em_v.fill_field(4, 62, ea_data['COV_PLAN_CODE'].strip().upper(), 6)
        em_v.fill_field(15, 20, ea_data['RATE_CD'].strip().upper(), 1)
        em_v.fill_field(5, 12, ea_data['NO_OF_UNITS'].strip().upper(), 3)
        em_v.fill_field(6, 71, ea_data['POL_EFF_DATE'].strip().upper(), 10)


        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_pf5()
        em_v.save_screen(htmlf_v)
        time.sleep(5)
        error_1 = em_v.string_get(23,3,38).strip()
        error_2 = em_v.string_get(23,42,38).strip()
      
        if em_v.string_get(2,18,37) == '*** BENEFIT ADD TRANSACTION (GSB) ***':   
            if(error_2 == 'DATABASE UPDATED'):
                print("GS trx successfully")
                em_v.send_pf2()
            elif (error_1 != ' ' or error_2 != ' '):  # Check for screen validation errors after submit
                print("GS-screen-error" + error_1 + error_2)
                raise ValueError("Screen validation failed. Please check the input fields.")


    except KeyError as ke:
        # Handle the case where 'orig1' key is missing from ea_data
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def policy_complete_add_screen_fill(ea_data, em_v, htmlf_v):
    try:
        # Fill fields using the provided data
        em_v.fill_field(24,10,'PR1 ',4)
        em_v.fill_field(24,20,'  ',2)
        em_v.fill_field(24,28,'T1',2)
        em_v.fill_field(24,38,ea_data['POL_NO'].strip().upper(),15)
        em_v.fill_field(24,61,' ',5)
        em_v.fill_field(24,70,ea_data['CO'].strip().upper(),3)
        em_v.fill_field(24,79,'N',1)

        
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)
        em_v.send_pf5()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(22,3,38)
        error_2 = em_v.string_get(22,42,38)
        error_3 = em_v.string_get(23,3,38)
        error_4 = em_v.string_get(23,42,38)
        
        if em_v.string_get(2,31,18) == '*** WELCOME TO ***' and em_v.string_get(24,10,4) == '    ':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("pr1-screen-error" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'any' key is missing from dict
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

def error_check(ea_data, em_v, htmlf_v):
    try:     
        # Save the screen and send enter
        em_v.save_screen(htmlf_v)
        em_v.send_enter()
        em_v.save_screen(htmlf_v)

        error_1 = em_v.string_get(21,1,80)
        error_2 = em_v.string_get(22,1,80)
        error_3 = em_v.string_get(23,1,80)
        error_4 = em_v.string_get(24,1,80)
        
        if em_v.string_get(2,35,10) == 'POLICY ADD' and em_v.string_get(20,50,4) == '    ':   
            if error_1 != ' ' or error_2 != ' ' or error_3 != ' '  or error_4 != ' ' :  # Check for screen validation errors after submit
                print("screen-error==>" + error_1 + error_2 + error_3 + error_4)
                raise ValueError("Screen validation failed. Please check the input fields.")
        

    except KeyError as ke:
        # Handle the case where 'any' key is missing from dict
        print(f"KeyError: Missing expected key in ea_data: {ke}")

    except AttributeError as ae:
        # Handle the case where methods like 'fill_field' are missing from em_v
        print(f"AttributeError: Missing attribute or method: {ae}")

    except CommandError as ce:
        # Handle the keyboard locked issue
        print(f"CommandError: {ce}. Keyboard is locked.")
        time.sleep(2)  # Wait for 2 seconds before retrying (adjust as necessary)
        
        # Retry sending the enter command or take necessary action to unlock the keyboard
        try:
            em_v.send_enter()  # Retry the enter command
        except CommandError as ce_retry:
            print(f"Retry failed due to CommandError: {ce_retry}. Exiting.")
            sys.exit(1)  # Exit if retry also fails
        
    except ValueError as ve:      
        print(f"Validation Error: {ve}") # Handle validation errors after sending enter
        sys.exit(1)  # Exit the program with an error status code

    except Exception as e:
        # Catch any other unexpected exceptions
        print(f"An unexpected error occurred: {e}")
        sys.exit(1)  # Exit the program with an error status code

